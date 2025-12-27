"""Excel文件处理工具"""
from typing import List, Dict, Any
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.schemas.member import MemberImportItem


# 成员导入模板列定义
MEMBER_TEMPLATE_COLUMNS = [
    ("学号", "student_id", True),
    ("姓名", "name", True),
    ("性别", "gender", False),
    ("寝室号", "dormitory", False),
    ("QQ邮箱", "qq_email", False),
]


def create_member_template() -> BytesIO:
    """创建成员导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成员导入模板"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, (header, _, required) in enumerate(MEMBER_TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header + ("*" if required else ""))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 15
    
    # 添加示例数据
    example_data = [
        ("2024001", "张三", "男", "A101", "123456789@qq.com"),
        ("2024002", "李四", "女", "B202", "987654321@qq.com"),
    ]
    
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_member_excel(file_content: bytes) -> List[MemberImportItem]:
    """解析成员Excel文件"""
    wb = load_workbook(filename=BytesIO(file_content))
    ws = wb.active
    
    members = []
    
    # 获取表头映射
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            # 移除必填标记
            header = header.replace("*", "").strip()
            for cn_name, en_name, _ in MEMBER_TEMPLATE_COLUMNS:
                if header == cn_name:
                    header_map[col] = en_name
                    break
    
    # 解析数据行
    for row in range(2, ws.max_row + 1):
        row_data = {}
        for col, field_name in header_map.items():
            value = ws.cell(row=row, column=col).value
            if value is not None:
                row_data[field_name] = str(value).strip()
        
        # 跳过空行
        if not row_data.get("student_id") or not row_data.get("name"):
            continue
        
        members.append(MemberImportItem(
            student_id=row_data.get("student_id", ""),
            name=row_data.get("name", ""),
            gender=row_data.get("gender"),
            dormitory=row_data.get("dormitory"),
            qq_email=row_data.get("qq_email"),
        ))
    
    return members


def export_members_to_excel(members: List[Dict[str, Any]]) -> BytesIO:
    """导出成员列表到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成员列表"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ["学号", "姓名", "性别", "寝室号", "QQ邮箱"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 15
    
    # 写入数据
    for row, member in enumerate(members, 2):
        ws.cell(row=row, column=1, value=member.get("student_id", "")).border = thin_border
        ws.cell(row=row, column=2, value=member.get("name", "")).border = thin_border
        ws.cell(row=row, column=3, value=member.get("gender", "")).border = thin_border
        ws.cell(row=row, column=4, value=member.get("dormitory", "")).border = thin_border
        ws.cell(row=row, column=5, value=member.get("qq_email", "")).border = thin_border
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
